#https://lektion-von-erfolglosigkeit.tistory.com/83

from openpyxl import load_workbook, Workbook

userName = 1
userId = 2
userMoney = 3
userLevel = 4

defaultMoney = 1000

wb = load_workbook("G:/내 드라이브/Coding/Python/userDataBase.xlsx")
ws = wb.active

def checkRow(): #빈 줄을 찾는 함수
    for row in range(2,ws.max_row+1):
        if ws.cell(row,1).value is None:
            return row

def signUp(_name,_id):
    _row = checkRow()

    ws.cell(row=_row, column=userName, value =_name)
    ws.cell(row=_row, column=userId, value=hex(_id)) #16진수로 저장
    ws.cell(row=_row, column=userMoney, value=defaultMoney)
    ws.cell(row=_row, column=userLevel, value=1)

    wb.save("G:/내 드라이브/Coding/Python/userDataBase.xlsx")

def checkName(_name,_id): #중복방지
    if ws.cell(2,1).value == None:
        return True
    else: 
        for row in range(2, ws.max_row+1):
            if ws.cell(row,1).value == _name and ws.cell(row,2).value == _id:
                return False
            else:
                return True

def userInfo(_name, _id):
    for row in range(2, ws.max_row+2):
    	if ws.cell(row, 1).value == _name and ws.cell(row, 2).value == _id:
        	 return ws.cell(row,userName),ws.cell(row,userMoney).value, ws.cell(row,userLevel).value

